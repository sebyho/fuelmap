#!/usr/bin/env python3
"""
Fetch national-average Euro-super 95 (petrol) & diesel prices for EU member
states from the European Commission's Weekly Oil Bulletin.

No API key required, but no clean JSON API either: the bulletin is
published as an .xlsx file. This script scrapes the bulletin's landing page
for the current "Prices with taxes latest prices (xlsx)" download link (the
underlying file is republished under a fresh URL each week, so we re-find it
every run rather than hardcoding one), then parses it with openpyxl.

Source page: https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en

Because this depends on the Commission not changing their sheet layout, the
parser is defensive: it hunts for header keywords ("EURO-SUPER"/"DIESEL"/
"GAS OIL") and country names rather than assuming fixed row/column numbers,
and it leaves the existing data/eu.json alone (rather than writing partial
or garbage data) if it can't find a reasonable number of countries.
"""

import json
import re
import sys
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from urllib.request import urlopen, Request
from urllib.error import HTTPError, URLError

try:
    from openpyxl import load_workbook
except ImportError:
    print("This script needs openpyxl: pip install openpyxl", file=sys.stderr)
    raise

ROOT = Path(__file__).resolve().parent.parent
GEO_FILE = ROOT / "scripts" / "geo_eu.json"
OUTPUT_FILE = ROOT / "data" / "eu.json"
BULLETIN_PAGE = "https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en"
MIN_COUNTRIES_EXPECTED = 15  # abort if we parse fewer than this — layout probably changed

HEADERS = {"User-Agent": "pump-fuelmap/1.0 (open-source fuel price map; https://github.com/sebyho/fuelmap)"}


def find_xlsx_url(html: str) -> str | None:
    # Look for the download link near the words "with Taxes" / "with taxes"
    matches = re.findall(r'href="([^"]+document/download/[^"]+\.xlsx[^"]*)"', html, re.IGNORECASE)
    for m in matches:
        if "tax" in m.lower():
            return m if m.startswith("http") else f"https://energy.ec.europa.eu{m}"
    return (matches[0] if matches[0].startswith("http") else f"https://energy.ec.europa.eu{matches[0]}") if matches else None


def fetch_bytes(url: str) -> bytes | None:
    try:
        req = Request(url, headers=HEADERS)
        with urlopen(req, timeout=30) as resp:
            return resp.read()
    except (HTTPError, URLError) as e:
        print(f"  request failed: {e}", file=sys.stderr)
        return None


def normalize(s) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip().upper()


def parse_workbook(raw_bytes: bytes, name_to_code: dict) -> dict:
    wb = load_workbook(BytesIO(raw_bytes), data_only=True)
    ws = wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))

    # 1. find the header row: the one containing both a petrol-ish and a
    #    diesel-ish keyword
    header_row_idx = None
    petrol_col = diesel_col = None
    for i, row in enumerate(rows[:20]):
        cells = [normalize(c) for c in row]
        p_col = next((j for j, c in enumerate(cells) if "SUPER" in c or "EURO-SUPER" in c or "EUROSUPER" in c or "UNLEADED" in c), None)
        d_col = next((j for j, c in enumerate(cells) if "DIESEL" in c or "GAS OIL" in c or "GASOIL" in c), None)
        if p_col is not None and d_col is not None:
            header_row_idx, petrol_col, diesel_col = i, p_col, d_col
            break

    if header_row_idx is None:
        raise ValueError("could not locate a header row with petrol + diesel columns")

    # 2. walk remaining rows, matching a country name in any of the first
    #    few cells and reading the identified petrol/diesel columns
    countries = {}
    for row in rows[header_row_idx + 1:]:
        if not row:
            continue
        country_code = None
        for cell in row[:3]:
            code = name_to_code.get(normalize(cell))
            if code and code not in ("EU", "EA"):
                country_code = code
                break
        if not country_code:
            continue

        def to_float(v):
            try:
                # The bulletin reports prices in EUR per 1000 litres, not
                # per litre — confirmed by the raw values being ~1000x a
                # plausible per-litre price (e.g. 2458.67 for Denmark, which
                # is right at its known ~€2.46/L level once converted).
                return round(float(v) / 1000, 3)
            except (TypeError, ValueError):
                return None

        petrol_val = to_float(row[petrol_col]) if petrol_col < len(row) else None
        diesel_val = to_float(row[diesel_col]) if diesel_col < len(row) else None
        if petrol_val or diesel_val:
            countries[country_code] = {"gasoline": petrol_val, "diesel": diesel_val}

    return countries


def main() -> int:
    geo = json.loads(GEO_FILE.read_text(encoding="utf-8"))

    print("Fetching Weekly Oil Bulletin page...")
    html_bytes = fetch_bytes(BULLETIN_PAGE)
    if not html_bytes:
        return 0
    html = html_bytes.decode("utf-8", errors="ignore")

    xlsx_url = find_xlsx_url(html)
    if not xlsx_url:
        print("Could not find the bulletin's xlsx download link — page layout may have changed.", file=sys.stderr)
        return 0

    print(f"Downloading {xlsx_url}")
    raw = fetch_bytes(xlsx_url)
    if not raw:
        return 0

    try:
        countries = parse_workbook(raw, geo["name_to_code"])
    except Exception as e:
        print(f"Failed to parse bulletin spreadsheet: {e}", file=sys.stderr)
        return 0

    if len(countries) < MIN_COUNTRIES_EXPECTED:
        print(
            f"Only parsed {len(countries)} countries (expected >= {MIN_COUNTRIES_EXPECTED}); "
            "leaving existing data/eu.json untouched.",
            file=sys.stderr,
        )
        return 0

    regions = []
    for code, prices in countries.items():
        centroid = geo["centroid"].get(code)
        if not centroid:
            continue
        entry = {
            "id": f"EU-{code}",
            "name": geo["code_to_display"].get(code, code),
            "level": "country",
            "lat": centroid[0],
            "lon": centroid[1],
        }
        if prices.get("gasoline"):
            entry["gasoline"] = {"value": prices["gasoline"]}
        if prices.get("diesel"):
            entry["diesel"] = {"value": prices["diesel"]}
        regions.append(entry)

    output = {
        "updated_at": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "source": "European Commission, Weekly Oil Bulletin (DG ENER)",
        "unit": "EUR/L",
        "note": "National average consumer prices including all taxes, as reported by each member state.",
        "regions": regions,
    }

    OUTPUT_FILE.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(regions)} countries to {OUTPUT_FILE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
